from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

ws.append(["번호", "영어", "수학"]) # A, B, C
for i in range(1, 11): 
  ws.append([i, randint(0, 100), randint(0, 100)])

col_B = ws["B"] # 영어 column 만 가지고 오기 
# print(col_B)
# for cell in col_B:
#   print(cell.value)
  
col_range = ws["B:C"] # 영어, 수학 column 함께 가지고 오기
# for cols in col_range:
#   for cell in cols:
#     print(cell.value)

row_title = ws[1] # 1번째 row 만 가지고 오기 
# for cell in row_title:
#   print(cell.value)

# row_range = ws[2:6] # 1(title)제외하고, 2번째 줄에서 "6번째"줄까지 가지고 온다. 
# for rows in row_range:
#   for cell in rows:
#     print(cell.value, end= " ")
#   print()

#################################################################
# from openpyxl.utils.cell import coordinate_from_string

# row_range = ws[2:ws.max_row] #* 2번째 줄부터 마지막 줄까지 
# for rows in row_range:
#   for cell in rows:
#     # print(cell.value, end=" ")
#     # print(cell.coordinate, end=" ") # A2 B2 C2 -> 각 cell 에대한 좌표 정보를 알 수 있음
#     xy = coordinate_from_string(cell.coordinate) 
#     print(xy, end=" ")
#     # print(xy[0], end="")
#     # print(xy[1], end=" ")
#   print()

# 전체 row 
# print(tuple(ws.rows)) # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.B1>, <Cell 'Sheet'.C1>)

# for row in tuple(ws.rows):
#   print(row[2].value)


# 전체 columns
# (<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>, <Cell 'Sheet'.A3>, <Cell 'Sheet'.A4>, <Cell 'Sheet'.A5>, <Cell 'Sheet'.A6>, <Cell 'Sheet'.A7>, <Cell 'Sheet'.A8>, <Cell 'Sheet'.A9>, <Cell 'Sheet'.A10>, <Cell 'Sheet'.A11>)
# print(tuple(ws.columns))

# for columns in tuple(ws.columns):
#   print(columns[0].value)

# for row in ws.iter_rows(): # 전체 row
#   print(row[0].value)

# for column in ws.iter_cols(): # 전체 column
#   print(column[0].value)

#! 범위를 지정한다.
# 1번째 줄부터 11번째 줄까지, 2번째 열부터 3번째 열까지
# for row in ws.iter_rows(min_row=1, max_row=11, min_col=2, max_col=3):
  # print(row[0].value, row[1].value) 
  # print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
  print(col) # (<Cell 'Sheet'.A1>, <Cell 'Sheet'.A2>, <Cell 'Sheet'.A3>, <Cell 'Sheet'.A4>, <Cell 'Sheet'.A5>)

wb.save("sample.xlsx")