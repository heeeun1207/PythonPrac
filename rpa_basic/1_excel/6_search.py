from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):  # 제목 행 제외
    # 번호, 영어, 수학
    if int(row[1].value) > 80:  # 영어 점수가 80 초과
        print(f"{row[0].value}번 학생은 영어 천재")

#* 과목명 수정이 필요할 때 
for row in ws.iter_rows(max_row=1): 
  for cell in row:
    if cell.value == "영어":
      cell.value = "컴퓨터"

wb.save("sample_modified.xlsx")