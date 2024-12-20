from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor = "ff9999"

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 2번째 index에 Sheet 생성

new_ws = wb["NewSheet"] # Dict 형태로 Sheet 에 접근 

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

print(wb.sheetnames)

wb.save("sample.xlsx")